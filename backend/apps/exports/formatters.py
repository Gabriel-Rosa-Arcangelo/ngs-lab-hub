from __future__ import annotations

import csv
import json
import os
import zipfile
from dataclasses import dataclass
from typing import Callable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from apps.datasets.models import LabResult

DATASET_LAB_RESULTS = "lab_results"
DATASET_NGS_TRIMMING = "ngs_trimming"
DATASET_NGS_PIPELINE = "ngs_pipeline"

LAB_RESULT_HEADERS = [
    "unit_code",
    "unit_name",
    "exam_code",
    "exam_name",
    "collected_at",
    "released_at",
    "tat_seconds",
    "status",
]

NGS_TRIMMING_HEADERS = [
    "sample_id",
    "platform",
    "pipeline_status",
    "raw_reads",
    "trimmed_reads",
    "trim_rate_percent",
    "q30_rate_percent",
    "read_length",
    "created_at",
]

NGS_PIPELINE_HEADERS = [
    "sample_id",
    "platform",
    "pipeline_status",
    "trimmed_reads",
    "aligned_reads",
    "alignment_rate_percent",
    "mean_depth",
    "variant_count",
    "created_at",
]

ProgressCallback = Callable[[int], None]


@dataclass(frozen=True)
class ExportBuildResult:
    file_path: str
    file_extension: str
    content_type: str
    row_count: int
    summary: dict


def build_export_file(
    file_format: str,
    queryset,
    temp_dir: str,
    dataset: str = DATASET_LAB_RESULTS,
    on_progress: ProgressCallback | None = None,
    progress_every: int = 2000,
) -> ExportBuildResult:
    if file_format == "csv":
        return build_csv_export(
            queryset,
            temp_dir,
            dataset=dataset,
            on_progress=on_progress,
            progress_every=progress_every,
        )
    if file_format == "xlsx":
        return build_xlsx_export(
            queryset,
            temp_dir,
            dataset=dataset,
            on_progress=on_progress,
            progress_every=progress_every,
        )
    if file_format == "zip":
        return build_zip_export(
            queryset,
            temp_dir,
            dataset=dataset,
            on_progress=on_progress,
            progress_every=progress_every,
        )
    raise ValueError(f"Unsupported export format: {file_format}")


def build_csv_export(
    queryset,
    temp_dir: str,
    dataset: str,
    on_progress: ProgressCallback | None = None,
    progress_every: int = 2000,
) -> ExportBuildResult:
    file_path = os.path.join(temp_dir, "results.csv")
    headers = _headers_for_dataset(dataset)

    with open(file_path, "w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(headers)
        row_count, summary = _write_rows(
            queryset,
            writer.writerow,
            dataset=dataset,
            on_progress=on_progress,
            progress_every=progress_every,
        )

    return ExportBuildResult(
        file_path=file_path,
        file_extension="csv",
        content_type="text/csv",
        row_count=row_count,
        summary=summary,
    )


def build_xlsx_export(
    queryset,
    temp_dir: str,
    dataset: str,
    on_progress: ProgressCallback | None = None,
    progress_every: int = 2000,
) -> ExportBuildResult:
    file_path = os.path.join(temp_dir, "results.xlsx")
    headers = _headers_for_dataset(dataset)

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="Results")

    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(worksheet, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
        header_cells.append(cell)
    worksheet.append(header_cells)

    row_count, summary = _write_rows(
        queryset,
        worksheet.append,
        dataset=dataset,
        on_progress=on_progress,
        progress_every=progress_every,
    )
    workbook.save(file_path)
    workbook.close()

    return ExportBuildResult(
        file_path=file_path,
        file_extension="xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        row_count=row_count,
        summary=summary,
    )


def build_zip_export(
    queryset,
    temp_dir: str,
    dataset: str,
    on_progress: ProgressCallback | None = None,
    progress_every: int = 2000,
) -> ExportBuildResult:
    csv_export = build_csv_export(
        queryset,
        temp_dir,
        dataset=dataset,
        on_progress=on_progress,
        progress_every=progress_every,
    )
    file_path = os.path.join(temp_dir, "results.zip")

    with zipfile.ZipFile(file_path, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.write(csv_export.file_path, arcname="results.csv")
        zip_file.writestr("summary.json", json.dumps(csv_export.summary, indent=2))

    return ExportBuildResult(
        file_path=file_path,
        file_extension="zip",
        content_type="application/zip",
        row_count=csv_export.row_count,
        summary=csv_export.summary,
    )


def _headers_for_dataset(dataset: str) -> list[str]:
    if dataset == DATASET_NGS_TRIMMING:
        return NGS_TRIMMING_HEADERS
    if dataset == DATASET_NGS_PIPELINE:
        return NGS_PIPELINE_HEADERS
    return LAB_RESULT_HEADERS


def _write_rows(
    queryset,
    append_row: Callable[[list], None],
    dataset: str,
    on_progress: ProgressCallback | None = None,
    progress_every: int = 2000,
) -> tuple[int, dict]:
    row_count = 0

    if dataset == DATASET_NGS_TRIMMING:
        pipeline_counts: dict[str, int] = {}
        trim_rate_sum = 0.0
        q30_sum = 0.0

        for sequence in queryset.iterator(chunk_size=2000):
            trim_rate = _safe_rate(sequence.trimmed_reads, sequence.raw_reads)
            append_row(
                [
                    sequence.sample_id,
                    sequence.platform,
                    sequence.pipeline_status,
                    sequence.raw_reads,
                    sequence.trimmed_reads,
                    trim_rate,
                    round(sequence.q30_rate_percent, 2),
                    sequence.read_length,
                    sequence.created_at.isoformat(),
                ]
            )
            row_count += 1
            trim_rate_sum += trim_rate
            q30_sum += sequence.q30_rate_percent
            pipeline_counts[sequence.pipeline_status] = pipeline_counts.get(sequence.pipeline_status, 0) + 1

            if on_progress and row_count % progress_every == 0:
                on_progress(row_count)

        if on_progress:
            on_progress(row_count)

        return row_count, {
            "counts": {
                "total": row_count,
                "by_pipeline_status": pipeline_counts,
            },
            "avg_trim_rate_percent": round(trim_rate_sum / row_count, 2) if row_count else 0.0,
            "avg_q30_rate_percent": round(q30_sum / row_count, 2) if row_count else 0.0,
        }

    if dataset == DATASET_NGS_PIPELINE:
        pipeline_counts = {}
        alignment_sum = 0.0
        mean_depth_sum = 0.0
        total_variants = 0

        for sequence in queryset.iterator(chunk_size=2000):
            alignment_rate = _safe_rate(sequence.aligned_reads, sequence.trimmed_reads)
            append_row(
                [
                    sequence.sample_id,
                    sequence.platform,
                    sequence.pipeline_status,
                    sequence.trimmed_reads,
                    sequence.aligned_reads,
                    alignment_rate,
                    round(sequence.mean_depth, 2),
                    sequence.variant_count,
                    sequence.created_at.isoformat(),
                ]
            )
            row_count += 1
            alignment_sum += alignment_rate
            mean_depth_sum += sequence.mean_depth
            total_variants += sequence.variant_count
            pipeline_counts[sequence.pipeline_status] = pipeline_counts.get(sequence.pipeline_status, 0) + 1

            if on_progress and row_count % progress_every == 0:
                on_progress(row_count)

        if on_progress:
            on_progress(row_count)

        return row_count, {
            "counts": {
                "total": row_count,
                "by_pipeline_status": pipeline_counts,
            },
            "avg_alignment_rate_percent": round(alignment_sum / row_count, 2) if row_count else 0.0,
            "avg_mean_depth": round(mean_depth_sum / row_count, 2) if row_count else 0.0,
            "total_variants": total_variants,
        }

    tat_sum = 0
    status_counts: dict[str, int] = {}

    for result in queryset.iterator(chunk_size=2000):
        append_row(_result_row(result))
        row_count += 1
        tat_sum += result.tat_seconds
        status_counts[result.status] = status_counts.get(result.status, 0) + 1

        if on_progress and row_count % progress_every == 0:
            on_progress(row_count)

    if on_progress:
        on_progress(row_count)

    avg_tat = round(tat_sum / row_count, 2) if row_count else 0.0
    return row_count, {
        "counts": {
            "total": row_count,
            "by_status": status_counts,
        },
        "avg_tat_seconds": avg_tat,
    }


def _safe_rate(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator * 100.0) / denominator, 2)


def _result_row(result: LabResult) -> list:
    return [
        result.unit.code,
        result.unit.name,
        result.exam_type.code,
        result.exam_type.name,
        result.collected_at.isoformat(),
        result.released_at.isoformat(),
        result.tat_seconds,
        result.status,
    ]

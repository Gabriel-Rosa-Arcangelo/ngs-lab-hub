import io
import json
import zipfile
from datetime import timedelta
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from apps.datasets.models import ExamType, LabResult, LabUnit
from apps.exports.models import ExportJob
from apps.exports.tasks import generate_export

User = get_user_model()


class GenerateExportFormatsTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="export_user", password="pass1234")
        self.unit = LabUnit.objects.create(code="U001", name="Main Lab")
        self.exam = ExamType.objects.create(code="EX0001", name="Basic Panel")

        now = timezone.now()
        LabResult.objects.create(
            unit=self.unit,
            exam_type=self.exam,
            collected_at=now - timedelta(hours=2),
            released_at=now - timedelta(hours=1),
            tat_seconds=3600,
            status=LabResult.Status.OK,
            patient_age=30,
            patient_sex="F",
        )
        LabResult.objects.create(
            unit=self.unit,
            exam_type=self.exam,
            collected_at=now - timedelta(hours=4),
            released_at=now - timedelta(hours=2),
            tat_seconds=7200,
            status=LabResult.Status.DELAYED,
            patient_age=40,
            patient_sex="M",
        )

    def test_generate_export_csv_happy_path(self):
        job = ExportJob.objects.create(created_by=self.user, file_format=ExportJob.FileFormat.CSV, params={"days": 30})
        captured = {}

        def fake_put_file(bucket: str, key: str, file_path: str, content_type: str):
            with open(file_path, "rb") as file_obj:
                captured["data"] = file_obj.read()
            captured["bucket"] = bucket
            captured["key"] = key
            captured["content_type"] = content_type

        with patch("apps.exports.tasks.put_file", side_effect=fake_put_file):
            generate_export.apply(args=[str(job.id)], throw=True)

        job.refresh_from_db()
        self.assertEqual(job.status, ExportJob.Status.READY)
        self.assertEqual(job.row_count, 2)
        self.assertEqual(job.progress_current, 2)
        self.assertEqual(job.progress_total, 2)
        self.assertTrue(captured["key"].endswith(".csv"))
        self.assertEqual(captured["content_type"], "text/csv")
        csv_text = captured["data"].decode("utf-8")
        self.assertIn("unit_code,unit_name,exam_code,exam_name,collected_at,released_at,tat_seconds,status", csv_text)
        self.assertIn("U001,Main Lab,EX0001,Basic Panel", csv_text)

    def test_generate_export_xlsx_happy_path(self):
        job = ExportJob.objects.create(created_by=self.user, file_format=ExportJob.FileFormat.XLSX, params={"days": 30})
        captured = {}

        def fake_put_file(bucket: str, key: str, file_path: str, content_type: str):
            with open(file_path, "rb") as file_obj:
                captured["data"] = file_obj.read()
            captured["key"] = key
            captured["content_type"] = content_type

        with patch("apps.exports.tasks.put_file", side_effect=fake_put_file):
            generate_export.apply(args=[str(job.id)], throw=True)

        job.refresh_from_db()
        self.assertEqual(job.status, ExportJob.Status.READY)
        self.assertEqual(job.row_count, 2)
        self.assertTrue(captured["key"].endswith(".xlsx"))
        self.assertEqual(captured["content_type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        workbook = load_workbook(io.BytesIO(captured["data"]))
        worksheet = workbook["Results"]
        header = [cell.value for cell in worksheet[1]]
        self.assertEqual(
            header,
            ["unit_code", "unit_name", "exam_code", "exam_name", "collected_at", "released_at", "tat_seconds", "status"],
        )
        self.assertEqual(worksheet.max_row, 3)

    def test_generate_export_zip_happy_path(self):
        job = ExportJob.objects.create(created_by=self.user, file_format=ExportJob.FileFormat.ZIP, params={"days": 30})
        captured = {}

        def fake_put_file(bucket: str, key: str, file_path: str, content_type: str):
            with open(file_path, "rb") as file_obj:
                captured["data"] = file_obj.read()
            captured["key"] = key
            captured["content_type"] = content_type

        with patch("apps.exports.tasks.put_file", side_effect=fake_put_file):
            generate_export.apply(args=[str(job.id)], throw=True)

        job.refresh_from_db()
        self.assertEqual(job.status, ExportJob.Status.READY)
        self.assertEqual(job.row_count, 2)
        self.assertTrue(captured["key"].endswith(".zip"))
        self.assertEqual(captured["content_type"], "application/zip")

        with zipfile.ZipFile(io.BytesIO(captured["data"])) as archive:
            names = set(archive.namelist())
            self.assertIn("results.csv", names)
            self.assertIn("summary.json", names)
            csv_text = archive.read("results.csv").decode("utf-8")
            summary = json.loads(archive.read("summary.json").decode("utf-8"))

        self.assertIn("U001,Main Lab,EX0001,Basic Panel", csv_text)
        self.assertEqual(summary["counts"]["total"], 2)
        self.assertEqual(summary["counts"]["by_status"]["OK"], 1)
        self.assertEqual(summary["counts"]["by_status"]["DELAYED"], 1)
        self.assertEqual(summary["avg_tat_seconds"], 5400.0)
